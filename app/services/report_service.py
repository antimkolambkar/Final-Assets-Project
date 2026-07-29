import io
import csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.models.employee import Employee, AccountStatus
from app.models.asset import Asset, AssetStatus, AssetAssignmentHistory
from app.models.vendor import Vendor, VendorRepairTicket
from app.models.ticket import Ticket, TicketStatus
from app.models.audit import AuditLog

class ReportService:
    REPORT_TYPES = {
        'employee': 'Employee Report',
        'asset': 'Asset Report',
        'assigned_asset': 'Assigned Asset Report',
        'available_asset': 'Available Asset Report',
        'repair_asset': 'Repair Asset Report',
        'department': 'Department Asset & Employee Report',
        'vendor': 'Vendor & Repair Ticket Report',
        'ticket': 'Ticket Report',
        'open_ticket': 'Open Ticket Report',
        'closed_ticket': 'Closed Ticket Report',
        'monthly_asset': 'Monthly Asset Movement Report',
        'monthly_ticket': 'Monthly Ticket Summary Report',
        'blocked_employee': 'Blocked Employee Report',
        'disabled_employee': 'Disabled Employee Report',
        'offboarded_employee': 'Offboarding & Return Asset Report',
        'asset_history': 'Asset History & Replacement Log',
        'audit': 'Audit Log Report'
    }

    @staticmethod
    def fetch_report_data(report_type, start_date=None, end_date=None, department=None, vendor_id=None, employee_id=None):
        headers = []
        rows = []
        title = ReportService.REPORT_TYPES.get(report_type, 'Enterprise ITAM Report')

        if report_type == 'employee':
            headers = ['Employee ID', 'Name', 'Email', 'Department', 'Designation', 'Manager', 'Location', 'Status', 'Last Synced']
            query = Employee.query
            if department:
                query = query.filter_by(department=department)
            if employee_id:
                query = query.filter_by(id=employee_id)
            for emp in query.all():
                rows.append([emp.employee_id, emp.name, emp.email, emp.department or '-', emp.designation or '-', emp.manager or '-', emp.office_location or '-', emp.account_status, emp.last_synced_at.strftime('%Y-%m-%d %H:%M') if emp.last_synced_at else '-'])

        elif report_type == 'asset':
            headers = ['Asset ID', 'Brand', 'Model', 'Serial Number', 'Processor', 'RAM', 'SSD', 'Vendor', 'Status', 'Assigned User']
            query = Asset.query
            if vendor_id:
                query = query.filter_by(vendor_id=vendor_id)
            for ast in query.all():
                rows.append([ast.asset_id, ast.brand, ast.model, ast.serial_number, ast.processor, ast.ram, ast.ssd, ast.vendor.name if ast.vendor else '-', ast.status, ast.assigned_user_name])

        elif report_type == 'assigned_asset':
            headers = ['Asset ID', 'Brand & Model', 'Serial Number', 'Specs (RAM/SSD)', 'Assigned User', 'Employee ID', 'Department', 'Email', 'Assignment Date']
            query = Asset.query.filter_by(status=AssetStatus.ASSIGNED)
            if vendor_id:
                query = query.filter_by(vendor_id=vendor_id)
            for ast in query.all():
                emp = ast.assigned_employee
                if department and emp and emp.department != department:
                    continue
                if employee_id and emp and emp.id != employee_id:
                    continue
                rows.append([
                    ast.asset_id, f"{ast.brand} {ast.model}", ast.serial_number, f"{ast.ram} / {ast.ssd}",
                    emp.name if emp else '-', emp.employee_id if emp else '-', emp.department if emp else '-',
                    emp.email if emp else '-', ast.assignment_date.strftime('%Y-%m-%d') if ast.assignment_date else '-'
                ])

        elif report_type == 'available_asset':
            headers = ['Asset ID', 'Brand', 'Model', 'Serial Number', 'Processor', 'RAM', 'SSD', 'Vendor', 'Status']
            query = Asset.query.filter_by(status=AssetStatus.AVAILABLE)
            if vendor_id:
                query = query.filter_by(vendor_id=vendor_id)
            for ast in query.all():
                rows.append([ast.asset_id, ast.brand, ast.model, ast.serial_number, ast.processor, ast.ram, ast.ssd, ast.vendor.name if ast.vendor else '-', ast.status])

        elif report_type == 'repair_asset':
            headers = ['Asset ID', 'Brand & Model', 'Serial Number', 'Vendor', 'Repair Status', 'Sent Date', 'Returned Date', 'Notes']
            query = VendorRepairTicket.query
            for vrt in query.all():
                ast = vrt.asset
                if vendor_id and vrt.vendor_id != vendor_id:
                    continue
                rows.append([
                    ast.asset_id if ast else '-', f"{ast.brand} {ast.model}" if ast else '-', ast.serial_number if ast else '-',
                    vrt.vendor.name if vrt.vendor else '-', vrt.repair_status,
                    vrt.sent_date.strftime('%Y-%m-%d') if vrt.sent_date else '-',
                    vrt.returned_date.strftime('%Y-%m-%d') if vrt.returned_date else '-',
                    vrt.notes or '-'
                ])

        elif report_type == 'department':
            headers = ['Department', 'Total Employees', 'Active', 'Blocked/Disabled', 'Offboarded', 'Assigned Assets']
            depts = db.session.query(Employee.department).distinct().all()
            for (dept_name,) in depts:
                if not dept_name:
                    continue
                if department and dept_name != department:
                    continue
                emp_count = Employee.query.filter_by(department=dept_name).count()
                active = Employee.query.filter_by(department=dept_name, account_status=AccountStatus.ACTIVE).count()
                blocked_dis = Employee.query.filter(Employee.department==dept_name, Employee.account_status.in_([AccountStatus.BLOCKED, AccountStatus.DISABLED])).count()
                offboarded = Employee.query.filter_by(department=dept_name, account_status=AccountStatus.OFFBOARDED).count()
                
                # Assigned assets to employees in this dept
                assigned_count = Asset.query.join(Employee).filter(Employee.department == dept_name).count()
                rows.append([dept_name, emp_count, active, blocked_dis, offboarded, assigned_count])

        elif report_type == 'vendor':
            headers = ['Vendor Name', 'Contact Person', 'Email', 'Phone', 'Total Assets Supplied', 'Under Repair Tickets']
            query = Vendor.query
            if vendor_id:
                query = query.filter_by(id=vendor_id)
            for v in query.all():
                asset_cnt = v.assets.count()
                repair_cnt = v.repair_tickets.filter(VendorRepairTicket.repair_status.in_(['Sent', 'In Repair'])).count()
                rows.append([v.name, v.contact_person or '-', v.email or '-', v.phone or '-', asset_cnt, repair_cnt])

        elif report_type in ['ticket', 'open_ticket', 'closed_ticket']:
            headers = ['Ticket ID', 'Subject', 'Employee Name', 'Department', 'Category', 'Priority', 'Assigned Engineer', 'Status', 'Created Date', 'Closed Date', 'Resolution Time (Hrs)']
            query = Ticket.query
            if report_type == 'open_ticket':
                query = query.filter(Ticket.status.notin_([TicketStatus.CLOSED, TicketStatus.RESOLVED]))
            elif report_type == 'closed_ticket':
                query = query.filter(Ticket.status.in_([TicketStatus.CLOSED, TicketStatus.RESOLVED]))
            if department:
                query = query.filter_by(department=department)
            for tkt in query.all():
                emp = tkt.employee
                eng = tkt.assigned_engineer
                rows.append([
                    tkt.ticket_id, tkt.subject, emp.name if emp else '-', tkt.department or '-', tkt.category, tkt.priority,
                    eng.full_name if eng else 'Unassigned', tkt.status,
                    tkt.created_at.strftime('%Y-%m-%d %H:%M'),
                    tkt.closed_at.strftime('%Y-%m-%d %H:%M') if tkt.closed_at else '-',
                    tkt.resolution_time_hours if tkt.resolution_time_hours else '-'
                ])

        elif report_type in ['blocked_employee', 'disabled_employee', 'offboarded_employee']:
            status_map = {
                'blocked_employee': AccountStatus.BLOCKED,
                'disabled_employee': AccountStatus.DISABLED,
                'offboarded_employee': AccountStatus.OFFBOARDED
            }
            target_status = status_map[report_type]
            headers = ['Employee ID', 'Name', 'Email', 'Department', 'Designation', 'Office Location', 'Status', 'Last Synced Date']
            query = Employee.query.filter_by(account_status=target_status)
            if department:
                query = query.filter_by(department=department)
            for emp in query.all():
                rows.append([emp.employee_id, emp.name, emp.email, emp.department or '-', emp.designation or '-', emp.office_location or '-', emp.account_status, emp.last_synced_at.strftime('%Y-%m-%d')])

        elif report_type == 'asset_history':
            headers = ['Log ID', 'Asset ID', 'Action', 'Employee Name', 'Old Asset', 'New Asset', 'Notes / Replacement Reason', 'Performed By', 'Timestamp']
            query = AssetAssignmentHistory.query.order_by(AssetAssignmentHistory.timestamp.desc())
            for h in query.all():
                ast = h.asset
                old_a = h.old_asset.asset_id if h.old_asset else '-'
                new_a = h.new_asset.asset_id if h.new_asset else '-'
                rows.append([
                    h.id, ast.asset_id if ast else '-', h.action, h.employee_name or '-', old_a, new_a,
                    h.notes or h.replacement_reason or '-', h.performed_by or 'System', h.timestamp.strftime('%Y-%m-%d %H:%M')
                ])

        elif report_type == 'audit':
            headers = ['Audit ID', 'User', 'Role', 'Action', 'Entity Type', 'Entity ID', 'IP Address', 'Details', 'Timestamp']
            query = AuditLog.query.order_by(AuditLog.timestamp.desc())
            for a in query.all():
                rows.append([
                    a.id, a.user_name, a.user_role or '-', a.action, a.entity_type or '-', a.entity_id or '-', a.ip_address or '-', a.details or '-', a.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ])

        else: # Default fallback
            headers = ['ID', 'Details']
            rows = [[1, 'General Summary']]

        return title, headers, rows

    @staticmethod
    def generate_excel(report_type, start_date=None, end_date=None, department=None, vendor_id=None, employee_id=None):
        title, headers, rows = ReportService.fetch_report_data(report_type, start_date, end_date, department, vendor_id, employee_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type[:30]

        # Title Banner
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=f"Enterprise ITAM - {title}")
        title_cell.font = Font(name='Segoe UI', size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Subtitle Timestamp
        ws.cell(row=2, column=1, value=f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}")
        ws.cell(row=2, column=1).font = Font(name='Segoe UI', size=9, italic=True, color="64748B")
        ws.row_dimensions[2].height = 20

        # Headers
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[4].height = 28

        # Data Rows
        row_alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        for row_idx, r in enumerate(rows, 5):
            ws.row_dimensions[row_idx].height = 22
            for col_idx, val in enumerate(r, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Segoe UI', size=10)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = row_alt_fill

        # Auto column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_csv(report_type, start_date=None, end_date=None, department=None, vendor_id=None, employee_id=None):
        title, headers, rows = ReportService.fetch_report_data(report_type, start_date, end_date, department, vendor_id, employee_id)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"# Enterprise ITAM Report: {title}"])
        writer.writerow([f"# Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"])
        writer.writerow([])
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        return output.getvalue().encode('utf-8')

    @staticmethod
    def generate_pdf(report_type, start_date=None, end_date=None, department=None, vendor_id=None, employee_id=None):
        title, headers, rows = ReportService.fetch_report_data(report_type, start_date, end_date, department, vendor_id, employee_id)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=15
        )

        elements.append(Paragraph(f"Enterprise ITAM - {title}", title_style))
        elements.append(Paragraph(f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", subtitle_style))

        # Format table data with Paragraph for wrap text
        cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#334155'))
        header_cell_style = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontSize=9, leading=11, fontName='Helvetica-Bold', textColor=colors.white)

        table_data = [[Paragraph(h, header_cell_style) for h in headers]]
        for r in rows:
            table_data.append([Paragraph(str(val), cell_style) for val in r])

        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))

        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
